import openpyxl
from Cal import PythonLecture

# score_table.xlsx파일 불러와서 데이터를 받아온다.
filename = "score_table.xlsx"

wb_obj = openpyxl.load_workbook(filename)
sheet_obj = wb_obj.active

max_row = 11    #행의 개수
max_col = 5     #열의 개수
currentCell_obj = sheet_obj.cell(row=1, column=1)
Score_list = []     #1~10까지 점수들과 결석 횟수가 담긴 list

# score_table_final.xlsx 파일을 생성하고, 데이터를 처리한다.
final1 = sheet_obj['F1']
final1.value = "Final Score"
final2 = sheet_obj['G1']
final2.value = "Final Grade"

for i in range(2, max_row+1):       #score_table.xlsx 파일에서 '점수'와 'absence'만 가져온다.
    for j in range(2, max_col+1):
        currentCell_obj = sheet_obj.cell(row = i, column=j)
        Score_list.append(currentCell_obj.value)        # 셀의 value를 Score_list에 추가한다.
    # PythonLecuture클래스의 weighted_avg함수와 get_Alphabet함수를 사용하여 평균(가중치)점수와 알파벳 점수를 구한다.
    Cal = PythonLecture(Score_list[0], Score_list[1], Score_list[2], Score_list[3])
    FinalScore = Cal.weighted_avg()
    FinalGrade = Cal.get_Alphabet(FinalScore, Score_list[3])

    # 위에서 구한 평균점수와 알파벳 점수를 새로운 셀에 추가한다.
    newCell_obj = sheet_obj['F'+str(i)]
    newCell_obj.value = FinalScore
    newCell2_obj = sheet_obj['G'+str(i)]
    newCell2_obj.value = FinalGrade
    wb_obj.save("score_table_final.xlsx")

    Score_list = []     # 한 행을 처리한 뒤 Score_list를 초기화 한다.




